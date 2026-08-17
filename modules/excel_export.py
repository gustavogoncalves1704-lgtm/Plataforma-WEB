"""Exportação de resultados em Excel (.xlsx) — 2 abas completas e formatadas."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COR_HEADER = "1F4E79"
COR_VERDE = "C6EFCE"
COR_VERMELHO = "FFC7CE"
COR_FAVORAVEL = "D9EAD3"
COR_DESFAVORAVEL = "F4CCCC"
COR_LINHA_ALT = "F5F5F5"

_FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONTE_DADOS = Font(name="Calibri", size=10, color="333333")
_FONTE_VERDE = Font(name="Calibri", size=10, bold=True, color="006100")
_FONTE_VERMELHO = Font(name="Calibri", size=10, bold=True, color="9C0006")

_BORDA = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

_ALIGN_WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
_ALIGN_WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

def gerar_excel(resultados: list[dict]) -> bytes:
    """Gera arquivo Excel com 2 abas: Produtos + Histórico."""
    wb = Workbook()
    _aba_produtos(wb, resultados)
    wb.remove(wb.active)
    _aba_historico(wb, resultados)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def _tem_dados_reais(produtos: list[dict]) -> bool:
    if not produtos:
        return False
    marca = str(produtos[0].get("marca", ""))
    return marca not in ("", "N/D", "None", "nan")

def _aba_produtos(wb: Workbook, resultados: list[dict]) -> None:
    ws = wb.create_sheet("Análise de Produtos")

    cabecalhos = [
        "Código SIAD", "Especificação Técnica",
        "Produto", "Marca", "Modelo", "Preço",
        "Custo-Benefício (0-10)", "Conformidade",
        "Checklist (Atendidos/Total)",
        "Link do Catálogo", "Contato do Fabricante",
        "Parecer", "Justificativa",
        "Modelo Recomendado", "Alerta de Direcionamento",
    ]

    larguras = {
        1: 15, 2: 60, 3: 12, 4: 20, 5: 30, 6: 15,
        7: 12, 8: 22, 9: 20, 10: 40, 11: 35,
        12: 22, 13: 70, 14: 30, 15: 35,
    }

    fill_header = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    for col, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = _FONTE_HEADER
        cell.fill = fill_header
        cell.alignment = _ALIGN_WRAP_CENTER
        cell.border = _BORDA
    ws.row_dimensions[1].height = 38

    linha = 2
    for item in resultados:
        codigo = item.get("codigo_siad", "N/A")
        espec = item.get("especificacao", "N/D")
        parecer = item.get("parecer", {})
        produtos = item.get("produtos", [])

        conclusao = parecer.get("parecer", "N/D")
        justificativa = parecer.get("justificativa", "N/D")
        recomendado = f"{parecer.get('marca_recomendada', '')} — {parecer.get('modelo_recomendado', '')}"
        alerta = next(
            (p.get("alerta_direcionamento", "") for p in produtos if p.get("alerta_direcionamento")),
            "",
        )

        if _tem_dados_reais(produtos):
            for i, prod in enumerate(produtos[:3], 1):
                _escrever_linha_produto(
                    ws, linha, codigo, espec, prod, i,
                    conclusao if i == 1 else "",
                    justificativa if i == 1 else "",
                    recomendado if i == 1 else "",
                    alerta if i == 1 else "",
                )
                linha += 1
        else:
            _escrever_linha_produto(
                ws, linha, codigo, espec, {}, 1,
                conclusao, justificativa, recomendado, alerta,
            )
            linha += 1

    for col, larg in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = larg

    for row_idx in range(2, linha):
        ws.row_dimensions[row_idx].height = 80

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{linha - 1}"
    ws.sheet_view.showGridLines = False

def _escrever_linha_produto(ws, linha, codigo, espec, prod, num, conclusao, justif, recomendado, alerta):
    conforme = prod.get("conforme", False) if prod else False
    pct = prod.get("conformidade_percentual", 0) if prod else 0
    status = f"{'✅ CONFORME' if conforme else '❌ NÃO CONFORME'} ({pct}%)" if prod else "—"

    checklist = prod.get("checklist", []) if prod else []
    if checklist:
        atendidos = sum(1 for c in checklist if c.get("atendido"))
        checklist_txt = f"{atendidos}/{len(checklist)} requisitos"
    else:
        checklist_txt = "N/D"

    dados = [
        codigo,
        espec,
        f"Produto {num}" if prod else "—",
        prod.get("marca", "N/D") if prod else "N/D",
        prod.get("modelo", "N/D") if prod else "N/D",
        prod.get("preco", "N/D") if prod else "N/D",
        prod.get("custo_beneficio", 0) if prod else 0,
        status,
        checklist_txt,
        prod.get("link_catalogo", "N/D") if prod else "N/D",
        prod.get("contato_fabricante", "N/D") if prod else "N/D",
        conclusao,
        justif if justif and justif != "N/D" else "N/D",
        recomendado,
        alerta,
    ]

    for col, val in enumerate(dados, 1):
        cell = ws.cell(row=linha, column=col, value=val)
        cell.font = _FONTE_DADOS
        cell.border = _BORDA
        cell.alignment = _ALIGN_WRAP_TOP

        if col == 8 and prod:
            if conforme:
                cell.fill = PatternFill(start_color=COR_VERDE, end_color=COR_VERDE, fill_type="solid")
                cell.font = _FONTE_VERDE
            else:
                cell.fill = PatternFill(start_color=COR_VERMELHO, end_color=COR_VERMELHO, fill_type="solid")
                cell.font = _FONTE_VERMELHO

        if col == 12 and conclusao:
            if "FAVOR" in conclusao.upper():
                cell.fill = PatternFill(start_color=COR_FAVORAVEL, end_color=COR_FAVORAVEL, fill_type="solid")
                cell.font = _FONTE_VERDE
            else:
                cell.fill = PatternFill(start_color=COR_DESFAVORAVEL, end_color=COR_DESFAVORAVEL, fill_type="solid")
                cell.font = _FONTE_VERMELHO

        if linha % 2 == 0 and col not in (8, 12):
            cell.fill = PatternFill(start_color=COR_LINHA_ALT, end_color=COR_LINHA_ALT, fill_type="solid")

def _aba_historico(wb: Workbook, resultados: list[dict]) -> None:
    ws = wb.create_sheet("Histórico de Licitações")

    cabecalhos = [
        "Código SIAD", "Especificação",
        "Nº do Processo", "Órgão",
        "Marca Vencedora", "Modelo",
        "Valor Unitário", "Data",
        "Fonte", "Link do Processo",
    ]

    larguras = {
        1: 15, 2: 40, 3: 25, 4: 30, 5: 20, 6: 25,
        7: 15, 8: 15, 9: 12, 10: 40,
    }

    fill_header = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    for col, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = _FONTE_HEADER
        cell.fill = fill_header
        cell.alignment = _ALIGN_WRAP_CENTER
        cell.border = _BORDA
    ws.row_dimensions[1].height = 38

    linha = 2
    for item in resultados:
        codigo = item.get("codigo_siad", "N/A")
        espec = item.get("especificacao", "N/D")[:80]
        historico = item.get("historico", [])

        if not historico:
            dados = [codigo, espec, "Sem histórico", "", "", "", "", "", "", ""]
            for col, val in enumerate(dados, 1):
                cell = ws.cell(row=linha, column=col, value=val)
                cell.font = _FONTE_DADOS
                cell.border = _BORDA
                cell.alignment = _ALIGN_WRAP_TOP
            ws.row_dimensions[linha].height = 30
            linha += 1
            continue

        for h in historico:
            dados = [
                codigo, espec,
                h.get("processo", "N/D"),
                h.get("orgao", "N/D"),
                h.get("marca", "N/D"),
                h.get("modelo", "N/D"),
                h.get("valor_unitario", "N/D"),
                h.get("data", "N/D"),
                h.get("fonte", "N/D"),
                h.get("link_processo", "N/D"),
            ]
            for col, val in enumerate(dados, 1):
                cell = ws.cell(row=linha, column=col, value=val)
                cell.font = _FONTE_DADOS
                cell.border = _BORDA
                cell.alignment = _ALIGN_WRAP_TOP
                if linha % 2 == 0:
                    cell.fill = PatternFill(start_color=COR_LINHA_ALT, end_color=COR_LINHA_ALT, fill_type="solid")
            ws.row_dimensions[linha].height = 30
            linha += 1

    for col, larg in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = larg

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{linha - 1}"
    ws.sheet_view.showGridLines = False